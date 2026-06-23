#!/usr/bin/env python3
"""Import «История болезни» Excel (labs dynamics, meds history, symptoms/MRI)
into health.db + clinical_history.json.

Usage: python scripts/import_excel_history.py ~/Downloads/История_болезни_Самадова.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402

from backend import health_db  # noqa: E402
from backend.dashboard_export import export_dashboard_html  # noqa: E402
from backend.paths import active_profile_id, health_root, profile_dir  # noqa: E402

# Column header month -> ISO date (mid-month, dates are approximate per source file)
PERIOD_DATES = {
    "нояб 2024": "2024-11-15",
    "янв 2025": "2025-01-15",
    "янв 2026": "2026-01-15",
}

TEST_CODE_MAP = {
    "пролактин": "prolactin",
    "макропролактин": "macroprolactin",
    "фсг": "fsh",
    "лг": "lh",
    "гормон роста": "gh",
    "ифр-1": "igf1",
    "эстрадиол": "estradiol",
    "прогестерон": "progesterone",
    "тестостерон": "testosterone",
    "дэас": "dheas",
    "ттг": "tsh",
    "т4 свободный": "ft4",
    "инсулин": "insulin",
    "17-он-прогестерон": "ohp17",
}

MARKERS_RE = re.compile(r"[✓✔⚠️↑↓→]|\s+", re.UNICODE)


def clean_num(raw: str) -> float | None:
    """'74,07 ✓' -> 74.07; '2268 ⚠️' -> 2268.0; '—', 'не опред.' -> None."""
    s = MARKERS_RE.sub(" ", str(raw)).strip()
    s = s.replace(",", ".").strip()
    m = re.match(r"^-?\d+(\.\d+)?$", s)
    return float(s) if m else None


def parse_norm(raw: str | None) -> tuple[float | None, float | None]:
    if not raw:
        return None, None
    s = str(raw).replace(",", ".")
    m = re.search(r"(\d+(?:\.\d+)?)\s*[–—-]\s*(\d+(?:\.\d+)?)", s)
    if m:
        return float(m.group(1)), float(m.group(2))
    m = re.search(r"<\s*(\d+(?:\.\d+)?)", s)
    if m:
        return None, float(m.group(1))
    m = re.search(r">\s*(\d+(?:\.\d+)?)", s)
    if m:
        return float(m.group(1)), None
    return None, None


def flag_for(value: float, lo: float | None, hi: float | None) -> str:
    if lo is not None and value < lo:
        return "low"
    if hi is not None and value > hi:
        return "high"
    return "normal"


def split_name_unit(raw: str) -> tuple[str, str | None]:
    """'Пролактин (мМЕ/л)' -> ('Пролактин', 'мМЕ/л')."""
    m = re.match(r"^(.*?)\s*\(([^)]*)\)\s*$", raw.strip())
    if m and ("/" in m.group(2) or "мЕ" in m.group(2) or "мк" in m.group(2) or "г" in m.group(2)):
        return m.group(1).strip(), m.group(2).strip()
    return raw.strip(), None


def code_for(name: str) -> str:
    low = name.lower()
    for key, code in TEST_CODE_MAP.items():
        if low.startswith(key):
            return code
    return re.sub(r"[^a-z0-9]+", "_", low).strip("_") or "unknown"


def import_labs(ws, conn, source_file: str) -> int:
    # Header row 3: cols C..E hold period labels
    header = [c.value for c in ws[3]]
    period_cols: list[tuple[int, str]] = []
    for idx, cell in enumerate(header):
        if not cell:
            continue
        key = str(cell).split("\n")[0].strip().lower()
        if key in PERIOD_DATES:
            period_cols.append((idx, PERIOD_DATES[key]))

    now = datetime.utcnow().isoformat()
    inserted = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        name_raw = row[0]
        if not name_raw:
            continue
        # Section header rows have empty norm + value cells
        if all(row[i] in (None, "") for i in range(1, min(6, len(row)))):
            continue
        name, unit = split_name_unit(str(name_raw))
        code = code_for(name)
        lo, hi = parse_norm(row[1] if len(row) > 1 else None)
        note = str(row[7]).strip() if len(row) > 7 and row[7] else None

        for col_idx, sample_date in period_cols:
            raw = row[col_idx] if col_idx < len(row) else None
            if raw in (None, "", "—") or "не опред" in str(raw):
                continue
            value = clean_num(str(raw))
            value_text = None if value is not None else str(raw).strip()
            if value is None and not value_text:
                continue
            flag = flag_for(value, lo, hi) if value is not None else "normal"

            dup = conn.execute(
                "SELECT 1 FROM lab_results WHERE test_code=? AND sample_date=?",
                (code, sample_date),
            ).fetchone()
            if dup:
                continue
            conn.execute(
                """INSERT INTO lab_results(
                    test_code, test_name_ru, value, value_text, unit,
                    ref_low, ref_high, flag, sample_date, source_file, created_at
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (code, name, value, value_text, unit, lo, hi, flag, sample_date, source_file, now),
            )
            inserted += 1
    conn.commit()
    return inserted


STATUS_MAP = [
    ("ПРИНИМАЕТ", "active"),
    ("ПЛАНИРУЕТСЯ", "planned"),
    ("НЕ РЕКОМЕНДОВАН", "not_recommended"),
    ("ОТМЕНЁН", "stopped"),
]


def med_status(raw: str | None) -> str:
    s = (raw or "").upper()
    for needle, status in STATUS_MAP:
        if needle in s:
            return status
    return "active"


def current_dose(dose_raw: str | None, status_raw: str | None) -> str:
    """'25 мкг → 50 мкг натощак' -> '50 мкг натощак'; prefer dose figure in status."""
    status = (status_raw or "").replace("✅", "").replace("ПРИНИМАЕТ", "").strip()
    if status and any(ch.isdigit() for ch in status):
        return status
    dose = str(dose_raw or "").replace("\n", " ").strip()
    if "Текущее:" in dose:
        return dose.split("Текущее:")[-1].strip()
    if "→" in dose:
        dose = dose.split("→")[-1].strip()
    return dose


def import_meds(ws, conn) -> tuple[int, list[dict], list[dict]]:
    """Main meds -> medications table. Returns (updated, nutraceuticals, evaluated)."""
    section = ""
    updated = 0
    nutraceuticals: list[dict] = []
    evaluated: list[dict] = []
    now = datetime.utcnow().isoformat()

    for row in ws.iter_rows(min_row=3, values_only=True):
        name_raw = row[0]
        if not name_raw:
            continue
        rest_empty = all(row[i] in (None, "") for i in range(1, len(row)))
        if rest_empty:
            section = str(name_raw).strip().upper()
            continue

        name_full = str(name_raw).strip()
        purpose = str(row[1] or "").strip()
        started = str(row[2] or "").strip()
        dose = current_dose(str(row[3] or ""), str(row[4] or ""))
        status = med_status(str(row[4] or ""))
        note = str(row[5] or "").replace("\n", " ").strip()

        if "ОСНОВНЫЕ" in section:
            base_name = name_full.split("(")[0].strip()
            generic_m = re.search(r"\(([^)]+)\)", name_full)
            generic = generic_m.group(1).strip() if generic_m else None
            existing = conn.execute(
                "SELECT id, generic FROM medications WHERE name LIKE ?",
                (base_name + "%",),
            ).fetchone()
            if existing:
                conn.execute(
                    "UPDATE medications SET dose=?, purpose=?, status=?, notes=? WHERE id=?",
                    (dose, purpose, status, note or started, existing[0]),
                )
            else:
                conn.execute(
                    """INSERT INTO medications(name, generic, dose, purpose, status, notes, started_at)
                       VALUES (?,?,?,?,?,?,?)""",
                    (base_name, generic, dose, purpose, status, note or started, now),
                )
            updated += 1
        elif "НУТРИЦЕВТИК" in section:
            nutraceuticals.append({
                "name": name_full, "purpose": purpose, "dose": dose,
                "status": status, "note": note, "since": started,
            })
        else:  # ОТМЕНЁННЫЕ / ОЦЕНЁННЫЕ
            evaluated.append({
                "name": name_full, "purpose": purpose, "status": status, "note": note,
            })
    conn.commit()
    return updated, nutraceuticals, evaluated


def import_symptoms(ws) -> dict:
    """Symptoms/MRI sheet -> structured dict for clinical_history.json."""
    section = ""
    dynamics: list[dict] = []
    mri: list[dict] = []
    pending: list[str] = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        name_raw = row[0]
        if not name_raw:
            continue
        rest_empty = all(row[i] in (None, "") for i in range(1, len(row)))
        if rest_empty:
            section = str(name_raw).strip()
            continue
        cells = [str(c).replace("\n", " ").strip() if c else "" for c in row[:5]]
        name, before, mid, current, comment = (cells + [""] * 5)[:5]

        if "МРТ" in section.upper():
            mri.append({"finding_ru": name, "before": before, "current": current, "comment_ru": comment})
        elif "PENDING" in section.upper() or "ДИАГНОСТИКА" in section.upper():
            pending.append(f"{name} — {comment}" if comment else name)
        else:
            dynamics.append({
                "section_ru": section, "symptom_ru": name,
                "before_2024": before, "at_3_months": mid, "current_2026": current,
                "comment_ru": comment,
            })
    return {"dynamics": dynamics, "mri": mri, "pending": pending}


def merge_clinical_history(profile: str, nutraceuticals, evaluated, symptoms) -> Path:
    path = profile_dir(profile) / "clinical_history.json"
    history = json.loads(path.read_text(encoding="utf-8")) if path.exists() else {}

    if nutraceuticals:
        history["nutraceuticals"] = [
            f"{n['name']} — {n['dose']}" + (f" ({n['note']})" if n["note"] else "")
            for n in nutraceuticals
        ]
    if evaluated:
        history["supplements_evaluated"] = evaluated
    if symptoms["dynamics"]:
        history["symptoms_dynamics"] = symptoms["dynamics"]
        history["symptoms_active"] = sorted({
            d["symptom_ru"] for d in symptoms["dynamics"]
            if "⚠" in d["current_2026"] or ("Нарушен" in d["current_2026"])
        } | set(history.get("symptoms_active", [])))
    if symptoms["mri"]:
        history["imaging_detail"] = symptoms["mri"]
    if symptoms["pending"]:
        history["pending_diagnostics_critical"] = symptoms["pending"]
    history["source_excel"] = "История_болезни_Самадова.xlsx (импортирована)"

    path.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx", help="Path to История_болезни .xlsx")
    parser.add_argument("--profile", default=None)
    args = parser.parse_args()

    profile = args.profile or active_profile_id()
    src = Path(args.xlsx).expanduser().resolve()
    if not src.exists():
        print(f"File not found: {src}")
        sys.exit(1)

    # Archive a copy next to other imports
    dest = health_root(profile) / "imports" / src.name
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dest)

    wb = openpyxl.load_workbook(src, data_only=True)
    conn = health_db.connect(profile)

    labs_added = import_labs(wb["Динамика анализов"], conn, src.name)
    meds_updated, nutraceuticals, evaluated = import_meds(wb["Препараты"], conn)
    symptoms = import_symptoms(wb["Симптомы и МРТ"])
    conn.close()

    history_path = merge_clinical_history(profile, nutraceuticals, evaluated, symptoms)
    export_dashboard_html(profile)

    print(f"Labs inserted:        {labs_added}")
    print(f"Medications updated:  {meds_updated}")
    print(f"Nutraceuticals:       {len(nutraceuticals)}")
    print(f"Evaluated/cancelled:  {len(evaluated)}")
    print(f"Symptom rows:         {len(symptoms['dynamics'])} (+{len(symptoms['mri'])} MRI, {len(symptoms['pending'])} pending)")
    print(f"Clinical history:     {history_path}")
    print("Dashboard refreshed.")


if __name__ == "__main__":
    main()
